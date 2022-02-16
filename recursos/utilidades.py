from asyncore import write
from pulp import *  


def converter_lista(lista=list):
    int_lista = [int(x) for x in lista]
    int_lista.sort()
    str_lista = [str(x) for x in int_lista]
    return str_lista


def coletar_custo(lista_do_custo=list):    
    lista_custo = []
    for lista in lista_do_custo:
        for valor in lista:
            lista_custo.append(valor)
    return lista_custo


def obter_indice_x_custo(lista_custo, lista_variaveis):
    custo_matriz = {}
    contador = 0
    for x in lista_variaveis:
        custo_matriz[f'{x}'] = lista_custo[contador]
        contador += 1
    return custo_matriz


def obter_resultado_por_variavel(variaveis_modulo, variaveis_dos_indices):
    resultado_variaveis_simples = {}
    for v in variaveis_modulo:
        resultado_variaveis_simples[f'{v.name}'] = v.value()
    resultado_por_variavel = {}
    for key in variaveis_dos_indices:
        resultado_por_variavel[f'{key}'] = resultado_variaveis_simples[f'{key}']
    return resultado_por_variavel


def obter_produto_do_custo(resultado_por_variavel, custo_e_indice):
    resultado_do_produto = {}
    for key in resultado_por_variavel:
        resultado_do_produto[f'{key}'] = resultado_por_variavel[f'{key}'] * custo_e_indice[f'{key}']
    return resultado_do_produto


def obter_resultado_por_empresa(num_empresas, localizacoes_dos_indices_tratados, num_ofertantes):
    resultados = {}
    for i in range(num_empresas):
        resultados[f'Empresa {str(i+1)}'] = lpSum(localizacoes_dos_indices_tratados[i][j].value() for j in range(num_ofertantes))
    return resultados


def tratamento_beta(resultado_final):
    from openpyxl import Workbook
    from recursos.dados import custo_oferta


    resultado_por_variavel = resultado_final['Resultado Final']['Resultado por variável']
    resultado_do_produto = resultado_final['Resultado Final']['Resultado por variável multiplicado']
    resultado_por_empresa = resultado_final['Resultado Final']['Resultado por empresa']
    custo_total = resultado_final['Resultado Final']['Custo Total']

    wb = Workbook()
    nome_do_arquivo = 'Resultado Final.xlsx'

    # Variáveis
    w_sheet_1 = wb.active
    w_sheet_1.title = 'Variáveis'
    w_sheet_1.append(['-', 'Ofertante 01', 'Ofertante 02', 'Ofertante 03', 'Ofertante 04', 'Ofertante 05', 'Ofertante 06', 'Ofertante 07', 'Ofertante 08', 'Não recebimento'])
    
    contador_de_coluna = 0
    contador_empresa = 0
    for valor in resultado_do_produto.keys():
        if contador_de_coluna == 0:
            contador_empresa += 1
            lista_coluna = [f'Empresa 0{contador_empresa}']
            contador_de_coluna += 1
        lista_coluna.append(f'{valor}')
        contador_de_coluna += 1
        if contador_de_coluna == 10:
            w_sheet_1.append(lista_coluna)
            contador_de_coluna = 0
    
    # Resultado por variável
    w_sheet_2 = wb.create_sheet(title='Resultado por variável')
    w_sheet_2.append(['-', 'Ofertante 01', 'Ofertante 02', 'Ofertante 03', 'Ofertante 04', 'Ofertante 05', 'Ofertante 06', 'Ofertante 07', 'Ofertante 08', 'Não recebimento'])

    contador_de_coluna = 0
    contador_empresa = 0
    for valor in resultado_por_variavel.values():
        if contador_de_coluna == 0:
            contador_empresa += 1
            lista_coluna = [f'Empresa 0{contador_empresa}']
            contador_de_coluna += 1
        lista_coluna.append(f'{valor}')
        contador_de_coluna += 1
        if contador_de_coluna == 10:
            w_sheet_2.append(lista_coluna)
            contador_de_coluna = 0

    # Custos
    w_sheet_3 = wb.create_sheet(title='Custos')
    w_sheet_3.append(['-', 'Ofertante 01', 'Ofertante 02', 'Ofertante 03', 'Ofertante 04', 'Ofertante 05', 'Ofertante 06', 'Ofertante 07', 'Ofertante 08', 'Não recebimento'])
    
    contador_de_coluna = 0
    contador_empresa = 0
    for valor in resultado_do_produto.values():
        if contador_de_coluna == 0:
            contador_empresa += 1
            lista_coluna = [f'Empresa 0{contador_empresa}']
            contador_de_coluna += 1
        lista_coluna.append(f'{valor}')
        contador_de_coluna += 1
        if contador_de_coluna == 10:
            w_sheet_3.append(lista_coluna)
            contador_de_coluna = 0
        
    # Restrição de oferta
    w_sheet_4 = wb.create_sheet(title='Restrição de oferta')
    w_sheet_4.append(['Ofertante 01', 'Ofertante 02', 'Ofertante 03', 'Ofertante 04', 'Ofertante 05', 'Ofertante 06', 'Ofertante 07', 'Ofertante 08', 'Não recebimento'])
    lista_coluna = [valor for valor in custo_oferta]
    w_sheet_4.append(lista_coluna)

    # Restrição de demandas
    w_sheet_5 = wb.create_sheet(title='Restrição de demandas')
    w_sheet_5.append(['Empresa 01', 'Empresa 02', 'Empresa 03', 'Empresa 04', 'Empresa 05', 'Empresa 06', 'Empresa 07', 'Empresa 08', 'Empresa 09', 'Empresa 10', 'Empresa 11', 'Empresa 12'])
    lista_coluna = [str(valor) for valor in resultado_por_empresa.values()]
    w_sheet_5.append(lista_coluna)
    
    # Custo total
    w_sheet_6 = wb.create_sheet(title='Custo Total')
    w_sheet_6.append(['Custo total'])
    w_sheet_6.append([f'{custo_total}'])

    wb.save(filename=nome_do_arquivo)    
    pass

